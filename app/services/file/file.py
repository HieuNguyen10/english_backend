from fastapi import Depends, File, UploadFile, status
from fastapi.responses import StreamingResponse
from fastapi_sqlalchemy import db
from fastapi_jwt_auth import AuthJWT
from app.serializers.base import DataResponse

from app.services.base import CRUDBase
from sqlalchemy.orm import Session
from app.db.base import get_db

from app.models.word.word import Word
from app.models.lesson.lesson import Lesson
from app.models.lesson_word.lesson_word import LessonWord

from app.serializers.word.word import *
from app.serializers.lesson_word.lesson_word import *
from app.serializers.lesson.lesson import *

from app.services.word.word import WordService

from starlette.exceptions import HTTPException
from sqlalchemy import and_
import openpyxl
import io
import re


class FileService(object):
    __instance = None

    @staticmethod
    def check_format(string):
        # Biểu thức chính quy kiểm tra chuỗi bắt đầu bằng 'LS' và theo sau là 4 chữ số
        pattern = r'^LS\d{4}$'
        return re.match(pattern, string) is not None

    @staticmethod
    def upload_file(file: File, db: Session = Depends(get_db)):
        try:
            workbook = openpyxl.load_workbook(filename=io.BytesIO(file))
            # Mở sheet theo tên
            sheet_name = "Word"  # Tên của sheet mà bạn muốn mở
            if sheet_name in workbook.sheetnames:  # Kiểm tra xem sheet có tồn tại không
                sheet = workbook[sheet_name]
            else:
                raise HTTPException(status_code=422, detail="Sheet not found")
            d = 1
            try:
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    d += 1
                    word_code = row[0]
                    word = WordRequestWithLesson()
                    word.english = row[1].strip()
                    word.type = row[2]
                    word.pronunciation = "" if row[3] is None else row[3]
                    word.vietnamese = row[4]
                    word.word_code_lesson = word_code
                    word = WordService.creat_word_lessson(word, db)
                    if word is None:
                        raise HTTPException(
                            status_code=422, detail=f"Word code at line {d} is invalid")
            except Exception as exc:
                print(exc)
            return True
        except Exception as exc:
            print(exc)
            raise HTTPException(status_code=422, detail="Something went wrong")

    @staticmethod
    def export_file(db: Session = Depends(get_db)):
        try:
            # Tạo workbook và sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Words"
            sheet.append(["Lesson Title", "English", "Type",
                          "Pronunciation", "Vietnamese", ])

            # Truy vấn lấy dữ liệu từ các bảng Word, LessonWord, Lesson
            query = (
                db.query(Word.word_code, Word.english, Word.type,
                         Word.pronunciation, Word.vietnamese, Lesson.title)
                .join(LessonWord, Word.id == LessonWord.word_id)
                .join(Lesson, LessonWord.lesson_id == Lesson.id)
                .all()
            )

            # Duyệt qua từng bản ghi và thêm vào sheet
            for word_code, english, word_type, pronunciation, vietnamese, lesson_title in query:
                sheet.append([lesson_title, english, word_type,
                              pronunciation, vietnamese])

            # Tạo luồng byte để lưu workbook
            stream = io.BytesIO()
            workbook.save(stream)
            stream.seek(0)

            # Trả về file Excel dưới dạng StreamingResponse
            return StreamingResponse(
                content=stream,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=words.xlsx"}
            )
        except Exception as exc:
            raise HTTPException(status_code=422, detail="Something went wrong")
